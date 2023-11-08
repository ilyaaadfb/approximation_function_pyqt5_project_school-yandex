import sys
from PyQt5.uic import loadUi
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QTableWidgetItem, QMainWindow, QFileDialog, QButtonGroup, QDialog
from PyQt5.QtWidgets import QApplication, QMessageBox
from PyQt5.QtSql import QSqlDatabase, QSqlTableModel
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
import matplotlib.pyplot as plt
import numpy as np
from scipy.optimize import curve_fit
import openpyxl
import pandas as pd
import sqlite3
from PyQt5 import QtCore, QtGui, QtWidgets
from main_interface import Ui_Function_approximation




class MainWindow(QMainWindow, Ui_Function_approximation):
    # интерфейс
    def __init__(self):
        super(MainWindow, self).__init__()
        # loadUi("main_interface2.ui", self)
        self.setupUi(self)
        self.figure = plt.figure()
        self.canvas = FigureCanvas(plt.figure())
        self.toolbar = NavigationToolbar(self.canvas, self)
        self.verticalLayout.addWidget(self.toolbar)
        self.verticalLayout.addWidget(self.canvas)
        self.setLayout(self.verticalLayout)
        self.spin.setMaximum(10000)
        self.button_group = QButtonGroup()
        self.button_group.addButton(self.linear_radioButton)
        self.button_group.addButton(self.nonlinear_radioButton)
        self.con = sqlite3.connect("db.sqlite")
        self.change_button.clicked.connect(self.update_result)
        self.table_for_db.itemChanged.connect(self.item_changed)
        self.save_button.clicked.connect(self.save_results)
        self.del_button.clicked.connect(self.delete_elem)
        self.action_exit.triggered.connect(self.exit_action)
        self.action_reference.triggered.connect(self.dialog_reference)
        self.action_rus.triggered.connect(self.leng_ru)
        self.action_en.triggered.connect(self.leng_en)
        self.modified = {}
        self.titles = None
        self.tabWidget.setTabText(0, "MAIN")
        self.tabWidget.setTabText(1, "БД")
        self.db()
        self.NameProgramm()
        self.contact()

    # выход из программы
    def exit_action(self):
        sys.exit(app.exec_())

    # диалоговое окно о программе
    def dialog_reference(self):
        self.dlg = QMessageBox(self)
        self.dlg.setWindowTitle("O программе")
        self.setWindowIcon(QIcon('graph_ava.png'))
        self.dlg.setText("     This is a program for creating linear and nonlinear approximations of functions. "
                         "This program was developed in Python using a set of PyQt5 libraries to create a graphical "
                         "interface. Linear approximation is performed using the 'pyplot' module in the 'matplotlib' "
                         "library. Nonlinear approximation is performed using the 'curve_fit' method in the 'Scipy' "
                         "library. The program also uses a 'SQLite3' database.")
        button = self.dlg.exec()

        if button == QMessageBox.Ok:
            print("OK!")

    # русский язык
    def leng_ru(self):
        self.button_file.setText("Выбор файла")
        self.clear_button_table.setText("Очистить таблицу")
        self.linear_radioButton.setText("Линейная")
        self.nonlinear_radioButton.setText("Нелинейная")
        self.compare_checkBox.setText("Сравнить")
        self.run_button.setText("ПУСК")
        self.clear_button_browser.setText("Очисть браузер")
        self.del_button.setText("Удалить")
        self.save_button.setText("Сохранить")
        self.change_button.setText("Загрузить")
        self.comboBox.setItemText(0, "Квадратичная: 'y = a*x^2+b*x+c'")
        self.comboBox.setItemText(1, "Кубическая: 'y = a*x^3 + b*x^2 + c*x + d'")
        self.comboBox.setItemText(2, "Степенная:  'y = k*x^n'")
        self.comboBox.setItemText(3, "Экспоненциальная I типа: 'y = a*exp(b^x)'")
        self.comboBox.setItemText(4, "Экспоненциальная II типа: 'y = a*b^x'")
        self.comboBox.setItemText(5, "Логарифмическая: 'y = b + a*log(x)'")
        self.comboBox.setItemText(6, "Гиперболическая:  'y = b+a/x'")

    # английский язык
    def leng_en(self):
        self.button_file.setText("File selection")
        self.clear_button_table.setText("Clear table")
        self.linear_radioButton.setText("Linear")
        self.nonlinear_radioButton.setText("Nonlinear")
        self.compare_checkBox.setText("Compare")
        self.run_button.setText("RUN")
        self.clear_button_browser.setText("Clear")
        self.del_button.setText("Clear")
        self.save_button.setText("Save")
        self.change_button.setText("Download")
        self.comboBox.setItemText(0, "Quadratic_func: 'y = a*x^2+b*x+c'")
        self.comboBox.setItemText(1, "Qubic_func: 'y = a*x^3 + b*x^2 + c*x + d'")
        self.comboBox.setItemText(2, "Power_func:  'y = k*x^n'")
        self.comboBox.setItemText(3, "Exponential_type_1_func: 'y = a*exp(b^x)'")
        self.comboBox.setItemText(4, "Exponential_type_2_func: 'y = a*b^x'")
        self.comboBox.setItemText(5, "Logarithmic_func: 'y = b + a*log(x)'")
        self.comboBox.setItemText(6, "Hyperbolic_func:  'y = b+a/x'")

        # горячие клавиши

    # горячие клавиши
    def keyPressEvent(self, event):
        #
        if int(event.modifiers()) == (Qt.ControlModifier + Qt.AltModifier):
            if event.key() == Qt.Key_Q:
                self.exit_action()

        if event.key() == Qt.Key_F1:
            self.dialog_reference()

    # добавление значения в базу данных
    def adding_value_in_db(self):
        connection = sqlite3.connect('db.sqlite')
        cursor = connection.cursor()
        value_func = self.comboBox.currentText()
        count_points = self.spin.value()
        if self.linear_radioButton.isChecked():
            cursor.execute('''INSERT INTO inf_ab_approx(count_points, function) VALUES (?, ?)''',
                           (count_points, "Линейная: 'y = kx + b'"))
        else:
            cursor.execute('''INSERT INTO inf_ab_approx(count_points, function) VALUES (?, ?)''',
                           (count_points, value_func))
        connection.commit()
        connection.close()
        self.db()

    # удаление элемента из базы данных
    def delete_elem(self):
        if self.id_spin.value() == 0:
            ...
        else:
            valid = QMessageBox.question(
                self, '', "Вы действительно хотите удалить элемент",
                QMessageBox.Yes, QMessageBox.No)
        try:
            if valid == QMessageBox.Yes:
                cur = self.con.cursor()
                cur.execute('DELETE FROM inf_ab_approx WHERE id = ?', (self.id_spin.value(),))
                self.con.commit()
        except:
            ...
        self.db()

    # сохранение результата изменения в базу данных
    def save_results(self):
        if self.modified:
            cur = self.con.cursor()
            que = "UPDATE inf_ab_approx SET\n"
            que += ", ".join([f"{key}='{self.modified.get(key)}'"
                              for key in self.modified.keys()])
            que += "WHERE id = ?"
            print(que)
            cur.execute(que, (self.id_spin.text(),))
            self.con.commit()
            self.modified.clear()
            self.db()

    # отслеживание изменений
    def item_changed(self, item):
        self.modified[self.titles[item.column()]] = item.text()

    # загрузка результата в базу данных
    def update_result(self):
        cur = self.con.cursor()
        result = cur.execute("SELECT * FROM inf_ab_approx WHERE id=?",
                             (self.id_spin.text(),)).fetchall()
        self.table_for_db.setRowCount(1)
        if not result:
            self.statusBar().showMessage('Ничего не нашлось')
            return
        else:
            self.statusBar().showMessage(f"Нашлась запись")
        self.table_for_db.setColumnCount(len(result[0]))
        self.table_for_db.setColumnWidth(0, 40)
        self.table_for_db.setColumnWidth(1, 40)
        self.table_for_db.setColumnWidth(2, 300)
        self.titles = [description[0] for description in cur.description]
        for i, elem in enumerate(result):
            for j, val in enumerate(elem):
                self.table_for_db.setItem(i, j, QTableWidgetItem(str(val)))
        self.modified = {}

    # база данных
    def db(self):
        db = QSqlDatabase.addDatabase('QSQLITE')
        db.setDatabaseName('db.sqlite')
        db.open()
        model = QSqlTableModel(self, db)
        model.setTable('inf_ab_approx')
        model.select()
        self.view.setModel(model)
        self.view.setColumnWidth(0, 40)
        self.view.setColumnWidth(1, 40)
        self.view.setColumnWidth(2, 300)
        pd.options.display.max_colwidth = 255

    # название и аватарка программы
    def NameProgramm(self):
        self.setWindowTitle('Function approximation')
        self.setWindowIcon(QIcon('graph_ava.png'))

    # взаимодействие интерфейса с методами
    def contact(self):
        self.button_file.clicked.connect(self.choice_file)
        self.clear_button_browser.clicked.connect(self.clear_textBrowser)
        self.clear_button_table.clicked.connect(self.clear_table)
        self.run_button.clicked.connect(self.show_graphic)
        self.spin.valueChanged.connect(self.change)
        self.button_group.buttonClicked.connect(self.on_radio_button_clicked)

    # проверка на то какой чек бокс нажат возвращает True в случае линейной и False в случае нелинейной
    def on_radio_button_clicked(self):
        if self.linear_radioButton.isChecked():
            return True
        else:
            return False

    # получение значений из таблицы по X и по Y
    def get_Value_table(self):
        self.value_table_X = []
        self.value_table_Y = []
        if self.tableWidget.rowCount() != 0:
            for i in range(self.tableWidget.rowCount()):
                try:
                    self.value_table_X.append(int(self.tableWidget.item(i, 0).text()))
                    self.value_table_Y.append(int(self.tableWidget.item(i, 1).text()))
                except:
                    ...
            if not self.value_table_X or not self.value_table_Y:
                self.popup_action()
            else:
                self.textBrowser.append(f'X: {", ".join(list(map(str, self.value_table_X)))}')
                self.textBrowser.append(f'Y: {", ".join(list(map(str, self.value_table_Y)))}')
                self.textBrowser.append(' ' * 50)
        else:
            self.popup_action()

    # функция появления всплывающего окна "введите точки для построения графика"
    def popup_action(self):
        error = QMessageBox()
        error.setWindowTitle("Ошибка")
        error.setText("Введите точки X и Y для выполнения этого действия")
        error.setIcon(QMessageBox.Warning)
        error.setStandardButtons(QMessageBox.Cancel | QMessageBox.Ok)
        error.setDefaultButton(QMessageBox.Ok)
        error.setInformativeText("")
        error.setDetailedText("""Детали
        Чтобы выполнить это действие введите точки по X и по Y

        Нажмите Ok, чтобы закрыть всплывающее окно

        Нажмите Cancel, чтобы выйти из программы
        """)
        error.buttonClicked.connect(self.popup_action_error)
        error.exec_()

    # функционал кнопок всплывающего окна
    def popup_action_error(self, btn):
        if btn.text() == "Ok":
            print("OK")
        elif btn.text() == "Cancel":
            sys.exit(app.exec_())

    # добавление строчки
    def change(self):
        self.tableWidget.setRowCount(int(self.spin.text()))
        self.dataExtent()

    # получение количества заполненных строк в таблице(количество точек)
    def dataExtent(self):
        self.maxRow = self.maxCol = -1
        for row in range(self.tableWidget.rowCount()):
            for col in range(self.tableWidget.columnCount()):
                item = self.tableWidget.item(row, col)
                if not item or not item.text():
                    continue
                self.maxRow = row
        self.maxRow += 1

    # выбор файла и запись его пути в тест браузер
    def choice_file(self):
        try:
            filename, filetype = QFileDialog.getOpenFileName(self, "Выбрать файл", ".", "XLSX Files(*.xlsx)")
            self.textBrowser.append("Путь к файлу:")
            self.textBrowser.append(filename)
            self.textBrowser.append(' ' * 50)
            self.filename_for_chice = filename
            self.load_data()
            self.dataExtent()
            self.maxRow_in_spin()
        except:
            ...

    # запись количества строчек в спин
    def maxRow_in_spin(self):
        self.spin.setValue(self.maxRow)

    # запись данных из файла в таблицу
    def load_data(self):
        path = str(self.filename_for_chice)
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        self.tableWidget.setRowCount(sheet.max_row)
        self.tableWidget.setColumnCount(sheet.max_column)
        self.tableWidget.setColumnWidth(0, 120)
        self.tableWidget.setColumnWidth(1, 120)
        list_values = list(sheet.values)
        self.tableWidget.setHorizontalHeaderLabels(list_values[0])
        row_index = 0
        for value_tuple in list_values[1:]:
            col_index = 0
            for value in value_tuple:
                self.tableWidget.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1

    # очищение текст браузера
    def clear_textBrowser(self):
        self.textBrowser.clear()

    # очистка таблицы
    def clear_table(self):
        self.tableWidget.clear()
        self.spin.setValue(0)

    # очистка графика
    def clear_graph(self):
        plt.clf()
        self.canvas.draw()

    # метод, который показывает точки на графике и задаёт диапазон по x и по y
    def points(self):
        plt.scatter(self.value_table_X, self.value_table_Y, color='red', s=15)  # точки
        plt.xlim([min(self.value_table_X) - 2, max(self.value_table_X) + 2])
        plt.ylim([min(self.value_table_Y) - 2, max(self.value_table_Y) + 2])
        plt.xlabel('x')
        plt.ylabel('y')

    # линейная аппроксимация
    def linear_approx(self):
        x = np.array(self.value_table_X)
        y = np.array(self.value_table_Y)
        k = np.polyfit(x, y, 1)[0]
        b = np.polyfit(x, y, 1)[1]
        self.textBrowser.append(f'Коэффицент k: {k}')
        self.textBrowser.append(f'Коэффицент b: {b}')
        self.textBrowser.append(' ' * 50)
        d = []
        for i in self.value_table_X:
            y = k * i + b
            d.append(y)
        plt.plot(self.value_table_X, d, color='blue', label="Линейная: 'y = kx + b'")
        plt.legend(loc='best')

    # методы выполняющие нелинейную аппроксимацию в зависимости от выбранной функции
    def quadratic_func(self):
        x = np.array(self.value_table_X)
        y = np.array(self.value_table_Y)

        def func(x, a, b, c):
            return a * x ** 2 + b * x + c

        popt, _ = curve_fit(func, x, y)
        x = np.linspace(x.min(), x.max(), 100)
        plt.plot(x, func(x, *popt), color='green', label="Квадратичная: 'y = a*x^2+b*x+c'")
        plt.legend(loc='best')

    def qubic_func(self):
        x = np.array(self.value_table_X)
        y = np.array(self.value_table_Y)

        def func(x, a, b, c, d):
            return a * x ** 3 + b * x ** 2 + c * x + d

        popt, _ = curve_fit(func, x, y)
        x = np.linspace(x.min(), x.max(), 100)
        plt.plot(x, func(x, *popt), color='violet', label="Кубическая: 'y = a*x^3 + b*x^2 + c*x + d'")
        plt.legend(loc='best')

    def power_func(self):
        x = np.array(self.value_table_X)
        y = np.array(self.value_table_Y)

        def func(x, k, n):
            return k * x ** n

        popt, _ = curve_fit(func, x, y)
        x = np.linspace(x.min(), x.max(), 100)
        plt.plot(x, func(x, *popt), color='brown', label="Степенная:  'y = k*x^n'")
        plt.legend(loc='best')

    def exponential_type_1_func(self):
        x = np.array(self.value_table_X)
        y = np.array(self.value_table_Y)

        def func(x, a, b):
            return a * np.exp(b ** x)

        popt, _ = curve_fit(func, x, y)
        x = np.linspace(x.min(), x.max(), 100)
        plt.plot(x, func(x, *popt), color='black', label="Экспоненциальная I типа: 'y = a*exp(b^x)'")
        plt.legend(loc='best')

    def exponential_type_2_func(self):
        x = np.array(self.value_table_X)
        y = np.array(self.value_table_Y)

        def func(x, a, b):
            return a * b ** x

        popt, _ = curve_fit(func, x, y)
        x = np.linspace(x.min(), x.max(), 100)
        plt.plot(x, func(x, *popt), color='orange', label="Экспоненциальная II типа: 'y = a*b^x'")
        plt.legend(loc='best')

    def logarithmic_func(self):
        x = np.array(self.value_table_X)
        y = np.array(self.value_table_Y)

        def func(x, a, b):
            return b + a * np.log(x)

        popt, _ = curve_fit(func, x, y)
        x = np.linspace(x.min(), x.max(), 100)
        plt.plot(x, func(x, *popt), color="grey", label="Логарифмическая: 'y = b + a*log(x)'")
        plt.legend(loc='best')

    def hyperbolic_func(self):
        x = np.array(self.value_table_X)
        y = np.array(self.value_table_Y)

        def func(x, a, b):
            return b + a / x

        popt, _ = curve_fit(func, x, y)
        x = np.linspace(x.min(), x.max(), 100)
        plt.plot(x, func(x, *popt), color='purple', label="Гиперболическая:  'y = b+a/x'")
        plt.legend(loc='best')

    # реализация графика
    def show_graphic(self):
        if self.compare_checkBox.isChecked():
            ...
        else:
            self.clear_graph()
        self.get_Value_table()
        if not self.value_table_X or not self.value_table_Y:
            ...
        else:
            self.adding_value_in_db()
            self.points()
            # линейная аппроксимация
            if self.on_radio_button_clicked():
                self.linear_approx()
            # нелинейная аппроксимация
            else:
                if self.nonlinear_radioButton.isChecked():
                    if self.comboBox.currentText() == "Квадратичная: 'y = a*x^2+b*x+c'":
                        self.quadratic_func()
                    elif self.comboBox.currentText() == "Кубическая: 'y = a*x^3 + b*x^2 + c*x + d'":
                        self.qubic_func()
                    elif self.comboBox.currentText() == "Степенная:  'y = k*x^n'":
                        self.power_func()
                    elif self.comboBox.currentText() == "Экспоненциальная I типа: 'y = a*exp(b^x)'":
                        self.exponential_type_1_func()
                    elif self.comboBox.currentText() == "Экспоненциальная II типа: 'y = a*b^x'":
                        self.exponential_type_2_func()
                    elif self.comboBox.currentText() == "Логарифмическая: 'y = b + a*log(x)'":
                        self.logarithmic_func()
                    elif self.comboBox.currentText() == "Гиперболическая:  'y = b+a/x'":
                        self.hyperbolic_func()
                else:
                    ...
            self.canvas.draw()


# запуск
if __name__ == '__main__':
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())
